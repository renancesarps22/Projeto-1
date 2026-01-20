import io
import os
import zipfile
from datetime import date, timedelta, datetime

from openpyxl import load_workbook

import pandas as pd
import streamlit as st
import plotly.express as px

# Optional DB (Supabase/Neon/Postgres)
try:
    from sqlalchemy import create_engine, text as sql_text
except Exception:  # pragma: no cover
    create_engine = None
    sql_text = None

# PDF export
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

st.set_page_config(page_title="App Personal", layout="wide")

DEFAULT_XLSX_PATH = "APP PERSONAL.xlsx"  # keep in same folder on deploy
REGISTRO_PATH = "registro_treinos.csv"   # local persistence
AVALIACOES_DB_PATH = "avaliacoes_db.csv" # base editável (inicializa a partir do Excel)

# Optional DB URL (set in Streamlit secrets as DATABASE_URL)
DATABASE_URL = (os.getenv("DATABASE_URL") or (st.secrets.get("DATABASE_URL", None) if hasattr(st, "secrets") else None))

BACKUP_ZIP_NAME = "backup_app_personal.zip"

# Excel template (aba de ficha)
TEMPLATE_SHEET_NAME = "FICHA_TREINO"  # conforme seu arquivo
FICHA_TABLE_START_ROW = 4  # primeira linha com exercícios
FICHA_MAX_ROWS = 9         # quantidade de linhas disponíveis no modelo (4..12)
FICHA_OBS_CELL = "B15"     # início do campo de observações (mesclado)

# -------------------------------------------------
# Helpers
# -------------------------------------------------

def _safe_to_datetime(s):
    return pd.to_datetime(s, errors="coerce")


def _load_workbook(uploaded_file: bytes | None):
    """Return dict of DataFrames for all sheets."""
    if uploaded_file is not None:
        xls = pd.ExcelFile(io.BytesIO(uploaded_file))
    else:
        if not os.path.exists(DEFAULT_XLSX_PATH):
            raise FileNotFoundError(
                f"Arquivo '{DEFAULT_XLSX_PATH}' nao encontrado. "
                "Envie o Excel pelo seletor da barra lateral ou coloque o arquivo junto do app."
            )
        xls = pd.ExcelFile(DEFAULT_XLSX_PATH)

    return {name: xls.parse(name) for name in xls.sheet_names}


def _to_date_col(df: pd.DataFrame, col: str = "Data") -> pd.DataFrame:
    if col in df.columns:
        df[col] = _safe_to_datetime(df[col]).dt.date
    return df


def _read_csv(path: str) -> pd.DataFrame:
    if os.path.exists(path):
        return pd.read_csv(path)
    return pd.DataFrame()


def _write_csv(df: pd.DataFrame, path: str):
    df.to_csv(path, index=False)



def _ensure_id(df: pd.DataFrame) -> pd.DataFrame:
    """Guarantee an ID column (string)."""
    if "ID" not in df.columns:
        df["ID"] = None
    if "Nome" in df.columns and "Data" in df.columns:
        missing = df["ID"].isna() | (df["ID"].astype(str).str.strip() == "")
        if missing.any():
            # stable id by Nome+Data+index
            base = (
                df.loc[missing, "Nome"].astype(str).str.strip().fillna("")
                + "|"
                + pd.to_datetime(df.loc[missing, "Data"], errors="coerce").astype(str)
                + "|"
                + df.loc[missing].index.astype(str)
            )
            df.loc[missing, "ID"] = base.apply(lambda x: str(abs(hash(x))))
    df["ID"] = df["ID"].astype(str)
    return df


def _db_url():
    """Return DATABASE_URL if configured (Streamlit secrets or env)."""
    try:
        if hasattr(st, "secrets") and "DATABASE_URL" in st.secrets:
            return str(st.secrets["DATABASE_URL"]).strip()
    except Exception:
        pass
    return (os.getenv("DATABASE_URL") or "").strip()

def _db_enabled() -> bool:
    return bool(_db_url()) and create_engine is not None

@st.cache_resource
def _get_engine():
    return create_engine(_db_url(), pool_pre_ping=True)

def _db_init_tables():
    if not _db_enabled():
        return
    eng = _get_engine()
    with eng.begin() as con:
        con.execute(sql_text("""
        CREATE TABLE IF NOT EXISTS treinos (
            id TEXT PRIMARY KEY,
            data DATE,
            nome TEXT,
            grupo_muscular TEXT,
            exercicio TEXT,
            series REAL,
            repeticoes REAL,
            carga_kg REAL,
            observacoes TEXT
        );
        """))
        con.execute(sql_text("""
        CREATE TABLE IF NOT EXISTS avaliacoes (
            id TEXT PRIMARY KEY
        );
        """))

def _db_upsert_df(table: str, df: pd.DataFrame):
    """Simple upsert by deleting IDs then inserting."""
    if not _db_enabled():
        return
    if df.empty or "ID" not in df.columns:
        return
    eng = _get_engine()
    ids = df["ID"].astype(str).tolist()
    with eng.begin() as con:
        con.execute(sql_text(f"DELETE FROM {table} WHERE id = ANY(:ids)"), {"ids": ids})
    # pandas will map columns, ensure id column name is id for db
    df2 = df.copy()
    if table == "treinos":
        df2 = df2.rename(columns={
            "ID": "id",
            "Data": "data",
            "Nome": "nome",
            "Grupo muscular": "grupo_muscular",
            "Exercicio": "exercicio",
            "Series": "series",
            "Repeticoes": "repeticoes",
            "Carga (kg)": "carga_kg",
            "Observacoes": "observacoes",
        })
    df2.to_sql(table, eng, if_exists="append", index=False)

def _db_read_df(table: str) -> pd.DataFrame:
    if not _db_enabled():
        return pd.DataFrame()
    _db_init_tables()
    eng = _get_engine()
    q = f"SELECT * FROM {table}"
    df = pd.read_sql(q, eng)
    if table == "treinos" and not df.empty:
        df = df.rename(columns={
            "id": "ID",
            "data": "Data",
            "nome": "Nome",
            "grupo_muscular": "Grupo muscular",
            "exercicio": "Exercicio",
            "series": "Series",
            "repeticoes": "Repeticoes",
            "carga_kg": "Carga (kg)",
            "observacoes": "Observacoes",
        })
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce").dt.date
    return df

def _sql_write_replace(df: pd.DataFrame, table: str):
    """Replace total (util para backup/edit)."""
    if not _db_enabled():
        return
    _db_init_tables()
    eng = _get_engine()
    df2 = df.copy()
    if table == 'treinos' and not df2.empty:
        df2 = df2.rename(columns={
            'ID': 'id',
            'Data': 'data',
            'Nome': 'nome',
            'Grupo muscular': 'grupo_muscular',
            'Exercicio': 'exercicio',
            'Series': 'series',
            'Repeticoes': 'repeticoes',
            'Carga (kg)': 'carga_kg',
            'Observacoes': 'observacoes',
        })
    df2.to_sql(table, eng, if_exists='replace', index=False)


def _kpi_delta(series: pd.Series):
    s = pd.to_numeric(series, errors="coerce").dropna()
    if len(s) < 2:
        return None
    return float(s.iloc[-1] - s.iloc[0])


def _fmt_delta(value, unit=""):
    if value is None:
        return "-"
    sign = "+" if value > 0 else ""
    return f"{sign}{value:.2f}{unit}"


def _current_value(series: pd.Series):
    s = pd.to_numeric(series, errors="coerce").dropna()
    if len(s) < 1:
        return None
    return float(s.iloc[-1])


def _to_float_or_none(x, treat_zero_as_none: bool = False):
    try:
        if x is None:
            return None
        v = float(x)
        if treat_zero_as_none and abs(v) < 1e-12:
            return None
        return v
    except Exception:
        return None


def _calc_imc(peso: float | None, altura: float | None) -> float | None:
    if peso is None or altura is None:
        return None
    if altura <= 0:
        return None
    return peso / (altura ** 2)


def _classificacao_imc(imc: float | None) -> str | None:
    if imc is None:
        return None
    if imc < 18.5:
        return "Baixo peso"
    if imc < 25:
        return "Peso normal"
    if imc < 30:
        return "Sobrepeso"
    if imc < 35:
        return "Obesidade Grau I"
    if imc < 40:
        return "Obesidade Grau II"
    return "Obesidade Grau III"


def _calc_dc_jp3(
    sexo: str | None,
    idade: float | None,
    d_pe: float | None,
    d_ab: float | None,
    d_cx: float | None,
    d_si: float | None,
    d_tr: float | None,
) -> float | None:
    """Densidade corporal (Jackson & Pollock 3 dobras).

    Homem: Peitoral + Abdominal + Coxa
    Mulher: Tríceps + Supra-ilíaca + Coxa
    """
    if sexo not in {"Homem", "Mulher"}:
        return None
    if idade is None:
        return None

    if sexo == "Homem":
        vals = [d_pe, d_ab, d_cx]
        if any(v is None for v in vals):
            return None
        s = sum(vals)
        return 1.10938 - 0.0008267 * s + 0.0000016 * (s**2) - 0.0002574 * idade

    vals = [d_tr, d_si, d_cx]
    if any(v is None for v in vals):
        return None
    s = sum(vals)
    return 1.0994921 - 0.0009929 * s + 0.0000023 * (s**2) - 0.0001392 * idade


def _calc_gordura_siri(dc: float | None) -> float | None:
    if dc is None or dc <= 0:
        return None
    return 495 / dc - 450


def _calc_rcq(cc: float | None, cq: float | None) -> float | None:
    if cc is None or cq is None or cq <= 0:
        return None
    return cc / cq


def _classificacao_rcq(sexo: str | None, rcq: float | None) -> str | None:
    """Classificação de risco (OMS simplificada)."""
    if rcq is None or sexo not in {"Homem", "Mulher"}:
        return None
    if sexo == "Homem":
        if rcq < 0.90:
            return "Baixo Risco"
        if rcq < 1.00:
            return "Risco Moderado"
        return "Alto Risco"

    if rcq < 0.80:
        return "Baixo Risco"
    if rcq < 0.85:
        return "Risco Moderado"
    return "Alto Risco"


def _recompute_derived(row: dict, base_cols: list[str]) -> dict:
    """Recalcula campos derivados para uma linha de avaliação."""
    out = {c: row.get(c) for c in base_cols}

    sexo = out.get("Sexo")
    idade = _to_float_or_none(out.get("Idade"), treat_zero_as_none=True)
    peso = _to_float_or_none(out.get("Peso"), treat_zero_as_none=True)
    altura = _to_float_or_none(out.get("Altura"), treat_zero_as_none=True)

    d_pe = _to_float_or_none(out.get("D PE"), treat_zero_as_none=True)
    d_ab = _to_float_or_none(out.get("D AB"), treat_zero_as_none=True)
    d_cx = _to_float_or_none(out.get("D CX"), treat_zero_as_none=True)
    d_si = _to_float_or_none(out.get("D SI"), treat_zero_as_none=True)
    d_tr = _to_float_or_none(out.get("D TR"), treat_zero_as_none=True)

    cc = _to_float_or_none(out.get("CC"), treat_zero_as_none=True)
    cq = _to_float_or_none(out.get("CQ"), treat_zero_as_none=True)

    imc = _calc_imc(peso, altura)
    out["IMC"] = imc
    imc_cls = _classificacao_imc(imc)
    # Sempre gravar a classificação (algumas planilhas usam "Classificação",
    # outras "Classificação IMC"). Se a coluna não existir no Excel original,
    # ainda assim exibimos no app.
    out["Classificação"] = imc_cls
    if "Classificação IMC" in base_cols:
        out["Classificação IMC"] = imc_cls

    dc = _calc_dc_jp3(sexo, idade, d_pe, d_ab, d_cx, d_si, d_tr)
    if "DC" in base_cols:
        out["DC"] = dc

    gordura = _calc_gordura_siri(dc)
    if "G" in base_cols:
        out["G"] = gordura
    if "% Gordura" in base_cols and gordura is not None:
        out["% Gordura"] = gordura

    if "MM" in base_cols and gordura is not None:
        out["MM"] = 100 - gordura
    if "% Massa Magra" in base_cols and gordura is not None:
        out["% Massa Magra"] = 100 - gordura

    rcq = _calc_rcq(cc, cq)
    out["RCQ"] = rcq
    out["RISCO"] = _classificacao_rcq(sexo, rcq)

    return out


def _make_pdf_from_table(title: str, df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    x0, y = 40, h - 50
    c.setFont("Helvetica-Bold", 14)
    c.drawString(x0, y, title)
    y -= 25

    c.setFont("Helvetica", 9)
    cols = list(df.columns)
    col_w = [70, 80, 90, 140, 40, 55, 55, 80]
    col_w = col_w[: len(cols)]
    if len(col_w) < len(cols):
        col_w += [80] * (len(cols) - len(col_w))

    for i, col in enumerate(cols):
        c.drawString(x0 + sum(col_w[:i]), y, str(col)[:18])
    y -= 14

    for _, row in df.iterrows():
        if y < 60:
            c.showPage()
            y = h - 50
            c.setFont("Helvetica", 9)
        for i, col in enumerate(cols):
            val = row.get(col, "")
            c.drawString(x0 + sum(col_w[:i]), y, str(val)[:24])
        y -= 12

    c.showPage()
    c.save()
    return buf.getvalue()


def _pt_weekday(d: date) -> str:
    # Monday=0
    nomes = [
        "SEGUNDA-FEIRA",
        "TERÇA-FEIRA",
        "QUARTA-FEIRA",
        "QUINTA-FEIRA",
        "SEXTA-FEIRA",
        "SÁBADO",
        "DOMINGO",
    ]
    try:
        return nomes[d.weekday()]
    except Exception:
        return "TREINO"


def _find_template_sheet(wb):
    if TEMPLATE_SHEET_NAME in wb.sheetnames:
        return wb[TEMPLATE_SHEET_NAME]
    # fallback: primeira aba que contenha FICHA e TREINO no nome
    for s in wb.sheetnames:
        s_norm = str(s).upper().replace(" ", "_")
        if "FICHA" in s_norm and "TREINO" in s_norm:
            return wb[s]
    # fallback: primeira aba
    return wb[wb.sheetnames[0]]


def _fill_ficha_sheet(ws, nome: str, d: date, itens: pd.DataFrame, obs: str = ""):
    """Preenche uma aba no modelo FICHA_TREINO preservando formatação."""
    # Cabeçalho (no seu modelo: B1="Nome:", C1=Nome; D1="Data:", E1=Data)
    # Se escrevermos "Nome: <nome>" em B1, o valor antigo de C1 pode ficar e o texto aparece duplicado.
    # Portanto, preservamos os rótulos e preenchemos apenas os campos de valor.
    ws["B1"].value = "Nome:"
    ws["C1"].value = str(nome)
    ws["D1"].value = "Data:"
    # Mantém como data (não texto) para o Excel formatar e para fórmulas funcionarem
    ws["E1"].value = datetime(d.year, d.month, d.day)

    # Linha do dia da semana (B2:F2 é mesclado no modelo)
    ws["B2"].value = _pt_weekday(d)

    # Limpar linhas da tabela no modelo
    start = FICHA_TABLE_START_ROW
    for i in range(FICHA_MAX_ROWS):
        r = start + i
        for c in ["B", "C", "D", "E", "F"]:
            ws[f"{c}{r}"].value = None

    # Preencher tabela (limitar ao espaço disponível)
    if itens is None or itens.empty:
        pass
    else:
        cols_map = {
            "grupo": next((c for c in itens.columns if "grupo" in str(c).lower()), None),
            "ex": next((c for c in itens.columns if "exerc" in str(c).lower()), None),
            "series": next((c for c in itens.columns if "serie" in str(c).lower()), None),
            "reps": next((c for c in itens.columns if "rep" in str(c).lower()), None),
            "carga": next((c for c in itens.columns if "carga" in str(c).lower()), None),
        }
        for idx, row in itens.reset_index(drop=True).iterrows():
            if idx >= FICHA_MAX_ROWS:
                break
            r = start + idx
            ws[f"B{r}"].value = str(row.get(cols_map["grupo"], "")) if cols_map["grupo"] else ""
            ws[f"C{r}"].value = str(row.get(cols_map["ex"], "")) if cols_map["ex"] else ""
            ws[f"D{r}"].value = row.get(cols_map["series"], "") if cols_map["series"] else ""
            ws[f"E{r}"].value = row.get(cols_map["reps"], "") if cols_map["reps"] else ""
            ws[f"F{r}"].value = row.get(cols_map["carga"], "") if cols_map["carga"] else ""

    # Observações
    if obs:
        ws[FICHA_OBS_CELL].value = obs


def _export_ficha_excel_model(
    xlsx_bytes: bytes | None,
    nome: str,
    modo: str,
    data_base: date,
    df_dia: pd.DataFrame,
    df_semana: pd.DataFrame,
    obs: str = "",
) -> bytes:
    """Gera Excel usando o layout existente do arquivo (aba FICHA_TREINO).

    modo: "Dia" ou "Semana"
    df_dia: itens do dia
    df_semana: itens da semana inteira (com coluna Data)
    """
    # Carregar workbook (do upload ou do arquivo padrão)
    if xlsx_bytes:
        buf = io.BytesIO(xlsx_bytes)
        wb = load_workbook(buf)
    else:
        wb = load_workbook(DEFAULT_XLSX_PATH)

    template = _find_template_sheet(wb)

    def _copy_with_title(title: str):
        ws_new = wb.copy_worksheet(template)
        # título único
        base = title[:31]
        if base in wb.sheetnames:
            k = 2
            while f"{base}_{k}" in wb.sheetnames:
                k += 1
            base = f"{base}_{k}"[:31]
        ws_new.title = base
        return ws_new

    if str(modo).lower().startswith("dia"):
        ws = _copy_with_title(_pt_weekday(data_base))
        _fill_ficha_sheet(ws, nome, data_base, df_dia, obs=obs)
    else:
        # Semana (segunda a domingo) da data_base
        start = data_base - timedelta(days=data_base.weekday())
        end = start + timedelta(days=6)
        dfw = df_semana.copy() if df_semana is not None else pd.DataFrame()
        if not dfw.empty and "Data" in dfw.columns:
            dfw["Data"] = pd.to_datetime(dfw["Data"], errors="coerce").dt.date
            dfw = dfw[(dfw["Data"] >= start) & (dfw["Data"] <= end)]

        # cria uma aba por dia da semana (segunda a domingo).
        # Mesmo que não haja itens no dia, a aba é criada (modelo preenchido com cabeçalho),
        # para facilitar impressão/organização semanal.
        for i in range(7):
            d = start + timedelta(days=i)
            itens_d = dfw[dfw.get("Data") == d] if (not dfw.empty and "Data" in dfw.columns) else pd.DataFrame()
            ws = _copy_with_title(_pt_weekday(d))
            _fill_ficha_sheet(ws, nome, d, itens_d, obs="")

    # Remover template original para não ir no arquivo final
    try:
        wb.remove(template)
    except Exception:
        pass

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def _make_pdf_report(nome: str, periodo: tuple, kpis: dict, rcq_table: pd.DataFrame) -> bytes:
    """PDF simples: KPIs + RCQ/Risco (texto)."""
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    x0, y = 40, h - 50

    c.setFont("Helvetica-Bold", 16)
    c.drawString(x0, y, "Relatorio - App Personal")
    y -= 24

    c.setFont("Helvetica", 10)
    c.drawString(x0, y, f"Nome: {nome}")
    y -= 14
    c.drawString(x0, y, f"Periodo: {periodo[0]} a {periodo[1]}")
    y -= 20

    c.setFont("Helvetica-Bold", 12)
    c.drawString(x0, y, "KPIs")
    y -= 16
    c.setFont("Helvetica", 10)
    for label, (val, delta, unit) in kpis.items():
        c.drawString(x0, y, f"- {label}: {val}{unit} (variacao: {delta}{unit})")
        y -= 14
        if y < 90:
            c.showPage()
            y = h - 50
            c.setFont("Helvetica", 10)

    y -= 10
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x0, y, "RCQ e Risco")
    y -= 16
    c.setFont("Helvetica", 10)
    if rcq_table is None or rcq_table.empty:
        c.drawString(x0, y, "Sem dados de RCQ no periodo selecionado.")
    else:
        cols = [c for c in ["Data", "RCQ", "RISCO"] if c in rcq_table.columns]
        for _, row in rcq_table[cols].head(15).iterrows():
            c.drawString(x0, y, f"{row.get('Data','')}  RCQ: {row.get('RCQ','')}  Risco: {row.get('RISCO','')}")
            y -= 14
            if y < 70:
                c.showPage()
                y = h - 50
                c.setFont("Helvetica", 10)

    c.save()
    return buf.getvalue()



def _make_backup_zip(avaliacoes_df: pd.DataFrame, treinos_df: pd.DataFrame) -> bytes:
    """Gera um ZIP com CSV + Excel das bases atuais."""
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, 'w', compression=zipfile.ZIP_DEFLATED) as z:
        # CSVs
        z.writestr('avaliacoes.csv', avaliacoes_df.to_csv(index=False))
        z.writestr('treinos.csv', treinos_df.to_csv(index=False))
        # Excel
        xbuf = io.BytesIO()
        with pd.ExcelWriter(xbuf, engine='openpyxl') as writer:
            avaliacoes_df.to_excel(writer, index=False, sheet_name='AVALIACOES')
            treinos_df.to_excel(writer, index=False, sheet_name='TREINOS')
        z.writestr('backup_app_personal.xlsx', xbuf.getvalue())
    return zbuf.getvalue()



def _read_registro() -> pd.DataFrame:
    # Banco (se configurado)
    if _db_enabled():
        df = _db_read_df("treinos")
    else:
        df = _read_csv(REGISTRO_PATH)

    if df.empty:
        df = pd.DataFrame(
            columns=[
                "ID",
                "Data",
                "Nome",
                "Grupo muscular",
                "Exercicio",
                "Series",
                "Repeticoes",
                "Carga (kg)",
                "Observacoes",
            ]
        )

    # padroniza
    if "Data" in df.columns:
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce").dt.date
    df = _ensure_id(df)
    return df


def _save_registro(df: pd.DataFrame):
    df = df.copy()
    if "Data" in df.columns:
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce").dt.date
    df = _ensure_id(df)

    if _db_enabled():
        _db_init_tables()
        # replace total (simples e confiavel)
        _sql_write_replace(df, "treinos")
    else:
        _write_csv(df, REGISTRO_PATH)


def _append_registro(rows: list[dict]):
    df_old = _read_registro()
    df_new = pd.DataFrame(rows)
    # garantir colunas
    for col in df_old.columns:
        if col not in df_new.columns:
            df_new[col] = None
    if "ID" not in df_new.columns:
        df_new["ID"] = None
    df_all = pd.concat([df_old, df_new[df_old.columns]], ignore_index=True)
    _save_registro(df_all)


def _load_or_init_avaliacoes_db(avaliacao_xlsx: pd.DataFrame) -> pd.DataFrame:
    """DB editável: se não existir, inicializa com a planilha."""
    if os.path.exists(AVALIACOES_DB_PATH):
        db = pd.read_csv(AVALIACOES_DB_PATH)
        db = _to_date_col(db, "Data")
        db = _ensure_id(db)
        return db

    base = avaliacao_xlsx.copy()
    base = _to_date_col(base, "Data")
    base = _ensure_id(base)
    _write_csv(base, AVALIACOES_DB_PATH)
    return base


def _save_avaliacoes_db(df: pd.DataFrame):
    df = df.copy()
    if "Data" in df.columns:
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce").dt.date
    _write_csv(df, AVALIACOES_DB_PATH)


# -------------------------------------------------
# Sidebar (Tema, Login opcional, Excel, filtros)
# -------------------------------------------------

st.sidebar.title("App Personal")

# Login opcional por senha (secrets / env). Se não definido, fica liberado.
with st.sidebar.expander("Acesso (opcional)"):
    senha_req = os.getenv("APP_PASSWORD") or st.secrets.get("APP_PASSWORD", None) if hasattr(st, "secrets") else None
    senha_in = st.text_input("Senha (se estiver configurada)", type="password")
    if senha_req:
        if senha_in != senha_req:
            st.sidebar.warning("Senha necessária para usar o app.")
    
# Tema
tema = st.sidebar.selectbox(
    "Tema do dashboard",
    ["Escuro (padrão)", "Claro", "Azul", "Verde"],
    index=0,
)

THEMES_CSS = {
    "Escuro (padrão)": "",
    "Claro": """
        <style>
        .stApp { background: #ffffff; }
        body, p, span, div, label, h1, h2, h3, h4, h5, h6 { color: #111111; }
        [data-testid="stHeader"] { background: rgba(255,255,255,0.90); }
        [data-testid="stMetricLabel"], [data-testid="stMetricValue"] { color: #111111 !important; }
        [data-testid="stDataFrame"] * { color: #111111 !important; }
        /* links */
        a { color: #0b57d0 !important; }
        </style>
    """,
    "Azul": """
        <style>
        .stApp { background: #071526; }
        [data-testid="stHeader"] { background: rgba(7,21,38,0.6); }
        </style>
    """,
    "Verde": """
        <style>
        .stApp { background: #071f16; }
        [data-testid="stHeader"] { background: rgba(7,31,22,0.6); }
        </style>
    """,
}

PLOTLY_TEMPLATE = "plotly_white" if tema == "Claro" else "plotly_dark"

st.markdown(THEMES_CSS.get(tema, ""), unsafe_allow_html=True)

# Layout compacto (bom para celular)
modo_mobile = st.sidebar.toggle("Modo celular (layout compacto)", value=st.session_state.get("modo_mobile", False))
st.session_state["modo_mobile"] = modo_mobile

# Modo aluno (somente leitura)
# - Configure FORCE_ALUNO=true em Secrets/Env para forçar modo aluno
force_aluno = False

# Link de aluno (somente leitura): adicione ?aluno=1 ao final da URL do app
# Ex.: https://seu-app.streamlit.app/?aluno=1
def _get_query_param(name: str):
    """Compat: st.query_params (novo) ou st.experimental_get_query_params (antigo)."""
    try:
        qp = st.query_params  # Streamlit >= 1.30
        val = qp.get(name)
        if isinstance(val, list):
            return val[0] if val else None
        return val
    except Exception:
        try:
            qp = st.experimental_get_query_params()
            val = qp.get(name)
            return val[0] if isinstance(val, list) and val else None
        except Exception:
            return None

def _truthy(v) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"1", "true", "t", "yes", "y", "sim", "on"}

# Se o link tiver ?aluno=1, forçamos modo aluno para esta sessão (sem precisar mexer no sidebar)
force_aluno_link = _truthy(_get_query_param("aluno"))
try:
    force_aluno = bool(st.secrets.get("FORCE_ALUNO", False))
except Exception:
    force_aluno = False
if os.getenv("FORCE_ALUNO") in {"1", "true", "True", "YES", "yes"}:
    force_aluno = True

# Aplicar força via link (secrets/env ainda têm prioridade, mas o link ativa quando não há secrets)
if force_aluno_link:
    force_aluno = True

if force_aluno:
    IS_STUDENT = True
else:
    IS_STUDENT = st.sidebar.toggle(
        "Modo aluno (somente leitura)",
        value=st.session_state.get("IS_STUDENT", False),
        help="Quando ligado, esconde adicionar/editar/excluir. Para forçar, use FORCE_ALUNO em Secrets.",
    )
    st.session_state["IS_STUDENT"] = IS_STUDENT


with st.sidebar.expander("Backup e exportacao"):
    st.caption("Baixe um backup das bases (avaliacoes + treinos) ja com edicoes/exclusoes.")
    try:
        _bk = _make_backup_zip(avaliacao_db if "avaliacao_db" in locals() else pd.DataFrame(), _read_registro())
        st.download_button("Baixar backup (ZIP)", data=_bk, file_name=BACKUP_ZIP_NAME, mime="application/zip", disabled=IS_STUDENT)
    except Exception as _e:
        st.info("Backup ficara disponivel apos carregar os dados.")

with st.sidebar.expander("Banco de dados (opcional)"):
    if _db_enabled():
        st.success("Banco configurado: usando DATABASE_URL")
    else:
        st.warning("Sem banco configurado: usando arquivos CSV locais")
    st.caption("Se quiser usar Supabase/Neon/Postgres, configure DATABASE_URL em Settings > Secrets no Streamlit Cloud.")

if IS_STUDENT:
    uploaded = None
else:
    uploaded = st.sidebar.file_uploader(
        "Envie o Excel (APP PERSONAL.xlsx)",
        type=["xlsx"],
        help="Se você não enviar, o app tenta carregar o arquivo padrão (APP PERSONAL.xlsx).",
    )

try:
    sheets = _load_workbook(uploaded.getvalue() if uploaded else None)
except Exception as e:
    st.error(str(e))
    st.stop()

avaliacao_xlsx = sheets.get("AVALIACAO_FISICA", pd.DataFrame()).copy()
dados_treinos = sheets.get("DADOS_TREINOS", pd.DataFrame()).copy()

if avaliacao_xlsx.empty:
    st.error("A aba 'AVALIACAO_FISICA' não foi encontrada ou está vazia.")
    st.stop()

# DB editável
avaliacao_db = _load_or_init_avaliacoes_db(avaliacao_xlsx)

# Se o usuário subir um Excel diferente e quiser reiniciar a base
with st.sidebar.expander("Configurações avançadas"):
    if (not IS_STUDENT) and st.button("Reinicializar avaliações com o Excel enviado"):
        base = _ensure_id(_to_date_col(avaliacao_xlsx.copy(), "Data"))
        _save_avaliacoes_db(base)
        st.success("Base de avaliações reinicializada.")
        st.rerun()

# filtros
avaliacao_db = _to_date_col(avaliacao_db, "Data")

nomes = sorted([x for x in avaliacao_db.get("Nome", pd.Series(dtype=str)).dropna().astype(str).unique().tolist() if x.strip()])

nome_sel = st.sidebar.selectbox("Nome", nomes if nomes else ["(sem nomes)"])

# Período rápido
if avaliacao_db["Data"].notna().any():
    dmax = avaliacao_db["Data"].max()
    dmin = avaliacao_db["Data"].min()
else:
    dmax = date.today()
    dmin = dmax

periodo_rapido = st.sidebar.selectbox(
    "Período rápido",
    ["Personalizado", "Últimos 30 dias", "Últimos 60 dias", "Últimos 90 dias", "Ano atual", "Tudo"],
    index=0,
)

if periodo_rapido == "Últimos 30 dias":
    d_start, d_end = dmax - timedelta(days=30), dmax
elif periodo_rapido == "Últimos 60 dias":
    d_start, d_end = dmax - timedelta(days=60), dmax
elif periodo_rapido == "Últimos 90 dias":
    d_start, d_end = dmax - timedelta(days=90), dmax
elif periodo_rapido == "Ano atual":
    d_start, d_end = date(dmax.year, 1, 1), dmax
elif periodo_rapido == "Tudo":
    d_start, d_end = dmin, dmax
else:
    dr = st.sidebar.date_input("Período", (dmin, dmax), min_value=dmin, max_value=dmax)
    if isinstance(dr, tuple) and len(dr) == 2:
        d_start, d_end = dr
    else:
        d_start, d_end = dmin, dmax

# Aplicar filtros
av = avaliacao_db.copy()
if "Nome" in av.columns and nomes:
    av = av[av["Nome"] == nome_sel]
if "Data" in av.columns:
    av = av[(av["Data"] >= d_start) & (av["Data"] <= d_end)]

av = av.sort_values("Data")

# -------------------------------------------------
# Tabs
# -------------------------------------------------

page = st.radio("Menu", ["Dashboard", "Avaliação Física", "Ficha de treino", "Registro de treinos"], horizontal=not modo_mobile)

# -----------------
# Dashboard
# -----------------
if page == "Dashboard":
    st.subheader("Dashboard")

    if av.empty:
        st.info("Nenhum dado encontrado para os filtros atuais.")
    else:
        # escolher colunas
        col_peso = "Peso" if "Peso" in av.columns else None
        col_g = "G" if "G" in av.columns else ("% Gordura" if "% Gordura" in av.columns else None)
        col_mm = "MM" if "MM" in av.columns else ("% Massa Magra" if "% Massa Magra" in av.columns else None)

        k1, k2, k3, k4, k5 = st.columns(5)

        if col_peso:
            peso_atual = _current_value(av[col_peso])
            delta_peso = _kpi_delta(av[col_peso])
            k1.metric("Peso (atual)", f"{peso_atual:.2f} kg" if peso_atual is not None else "-", _fmt_delta(delta_peso, " kg"))
        else:
            k1.metric("Peso (atual)", "-")

        if col_g:
            g_atual = _current_value(av[col_g])
            delta_g = _kpi_delta(av[col_g])
            k2.metric("% Gordura (atual)", f"{g_atual:.2f} %" if g_atual is not None else "-", _fmt_delta(delta_g, " %"))
        else:
            k2.metric("% Gordura (atual)", "-")

        if col_mm:
            mm_atual = _current_value(av[col_mm])
            delta_mm = _kpi_delta(av[col_mm])
            k3.metric("% Massa magra (atual)", f"{mm_atual:.2f} %" if mm_atual is not None else "-", _fmt_delta(delta_mm, " %"))
        else:
            k3.metric("% Massa magra (atual)", "-")

        if "IMC" in av.columns:
            imc_atual = _current_value(av["IMC"])
            delta_imc = _kpi_delta(av["IMC"])
            k4.metric("IMC (atual)", f"{imc_atual:.2f}" if imc_atual is not None else "-", _fmt_delta(delta_imc, ""))
        else:
            k4.metric("IMC (atual)", "-")

        if "RCQ" in av.columns and av["RCQ"].notna().any():
            rcq_atual = float(pd.to_numeric(av["RCQ"], errors="coerce").dropna().iloc[-1])
            risco_atual = av["RISCO"].dropna().iloc[-1] if "RISCO" in av.columns and av["RISCO"].notna().any() else "-"
            k5.metric("RCQ (atual)", f"{rcq_atual:.2f}", str(risco_atual), delta_color="off")
        else:
            k5.metric("RCQ (atual)", "-")

        # Relatorio PDF (Dashboard)
        kpis_report = {
            'Peso': (f'{peso_atual:.2f}' if 'peso_atual' in locals() and peso_atual is not None else '-', _fmt_delta(delta_peso, ''), ' kg'),
            '% Gordura': (f'{g_atual:.2f}' if 'g_atual' in locals() and g_atual is not None else '-', _fmt_delta(delta_g, ''), ' %'),
            '% Massa magra': (f'{mm_atual:.2f}' if 'mm_atual' in locals() and mm_atual is not None else '-', _fmt_delta(delta_mm, ''), ' %'),
            'IMC': (f'{imc_atual:.2f}' if 'imc_atual' in locals() and imc_atual is not None else '-', _fmt_delta(delta_imc, ''), ''),
            'RCQ': (f'{rcq_atual:.2f}' if 'rcq_atual' in locals() and rcq_atual is not None else '-', '-', ''),
        }
        rcq_tbl = av[[c for c in ['Data','RCQ','RISCO'] if c in av.columns]].copy() if 'av' in locals() else pd.DataFrame()

        # Em alguns ambientes (ex.: Cloud em reruns), variáveis podem não existir por conta de execuções parciais.
        # Garanta um nome seguro para o relatório.
        nome_report = None
        nome_report = 'avaliacao'
        try:
            nome_report = str(nome_sel)  # sidebar
        except Exception:
            nome_report = None
        if not nome_report or nome_report == '(sem nomes)':
            if 'av' in locals() and isinstance(av, pd.DataFrame) and (not av.empty) and ('Nome' in av.columns):
                nome_report = str(av['Nome'].astype(str).iloc[-1])

        if not nome_report:
            nome_report = 'avaliacao'
        # Garantir datas para o relatório (evita NameError em data_i/data_f)
        from datetime import date as _date
        
        # Se o app já tiver definido data_i/data_f em outro ponto, respeite.
        # Caso contrário, inferimos do dataframe filtrado (av) ou usamos hoje.
        if ('data_i' not in locals()) or (data_i is None):
            try:
                if 'av' in locals() and isinstance(av, pd.DataFrame) and (not av.empty) and ('Data' in av.columns):
                    _min = pd.to_datetime(av['Data'], errors='coerce').dropna().min()
                    data_i = _min.date() if hasattr(_min, 'date') else _min
                else:
                    data_i = _date.today()
            except Exception:
                data_i = _date.today()
        
        if ('data_f' not in locals()) or (data_f is None):
            try:
                if 'av' in locals() and isinstance(av, pd.DataFrame) and (not av.empty) and ('Data' in av.columns):
                    _max = pd.to_datetime(av['Data'], errors='coerce').dropna().max()
                    data_f = _max.date() if hasattr(_max, 'date') else _max
                else:
                    data_f = _date.today()
            except Exception:
                data_f = _date.today()
        pdf_rel = _make_pdf_report(nome_report, (data_i, data_f), kpis_report, rcq_tbl)
        st.download_button('Exportar relatorio (PDF)', data=pdf_rel, file_name=f'relatorio_{nome_report}.pdf', mime='application/pdf')

        st.divider()

        c1, c2 = st.columns(2)
        with c1:
            with st.container(border=True):
                if col_peso:
                    fig = px.line(av, x="Data", y=col_peso, markers=True, title="Peso ao longo do tempo", template=PLOTLY_TEMPLATE)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("Coluna de peso não encontrada.")

            with st.container(border=True):
                if col_g:
                    fig = px.line(av, x="Data", y=col_g, markers=True, title="% Gordura ao longo do tempo", template=PLOTLY_TEMPLATE)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("Coluna de % gordura não encontrada.")

        with c2:
            with st.container(border=True):
                if col_mm:
                    fig = px.line(av, x="Data", y=col_mm, markers=True, title="% Massa magra ao longo do tempo", template=PLOTLY_TEMPLATE)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("Coluna de % massa magra não encontrada.")

            with st.container(border=True):
                circ_cols = [c for c in ["CC", "CQ", "CA"] if c in av.columns]
                if circ_cols:
                    df_melt = av[["Data"] + circ_cols].melt(id_vars=["Data"], var_name="Circunferência", value_name="Valor")
                    fig = px.line(df_melt, x="Data", y="Valor", color="Circunferência", markers=True, title="Circunferências ao longo do tempo", template=PLOTLY_TEMPLATE)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("Circunferências (CC/CQ/CA) não encontradas.")

        c3, c4 = st.columns(2)
        with c3:
            with st.container(border=True):
                if "RCQ" in av.columns:
                    fig = px.line(av, x="Data", y="RCQ", markers=True, title="RCQ ao longo do tempo", template=PLOTLY_TEMPLATE)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("Coluna RCQ não encontrada.")

        with c4:
            with st.container(border=True):
                st.markdown("#### Quadro RCQ e risco")
                cols = [c for c in ["Data", "RCQ", "RISCO"] if c in av.columns]
                st.dataframe(av[cols], use_container_width=True, hide_index=True)

# -----------------
# Avaliação Física
# -----------------
if page == "Avaliação Física":
    st.subheader("Avaliação Física")

    # tabela (respeita filtros)
    st.caption("A tabela abaixo respeita os filtros de Nome e Período da barra lateral.")

    # seleção múltipla para excluir
    av_view = av.copy()
    if not av_view.empty:
        av_view.insert(0, "Selecionar", False)
        edited = st.data_editor(
            av_view,
            use_container_width=True,
            hide_index=True,
            disabled=[c for c in av_view.columns if c not in {"Selecionar"}],
            key="editor_av",
        )
        selected_ids = edited.loc[edited["Selecionar"] == True, "ID"].astype(str).tolist()

        b1, b2, b3 = st.columns([1, 1, 2])
        with b1:
            if st.button("Excluir selecionadas", disabled=(IS_STUDENT or len(selected_ids) == 0)):
                st.session_state["confirm_delete"] = True
        with b2:
            if st.button("Exportar base atual (Excel)"):
                # exporta todas as avaliações do DB (não só filtradas)
                db = avaliacao_db.copy()
                out = io.BytesIO()
                with pd.ExcelWriter(out, engine="openpyxl") as writer:
                    db.to_excel(writer, index=False, sheet_name="AVALIACAO_FISICA")
                st.download_button(
                    "Baixar Excel (base completa)",
                    data=out.getvalue(),
                    file_name="avaliacoes_atualizadas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        with b3:
            st.write("")

        if st.session_state.get("confirm_delete"):
            st.warning(f"Confirmar exclusão de {len(selected_ids)} avaliação(ões)?")
            c1, c2 = st.columns(2)
            with c1:
                if st.button("✅ Confirmar exclusão"):
                    db = avaliacao_db.copy()
                    db = db[~db["ID"].astype(str).isin([str(x) for x in selected_ids])]
                    _save_avaliacoes_db(db)
                    st.success("Avaliações excluídas.")
                    st.session_state["confirm_delete"] = False
                    st.rerun()
            with c2:
                if st.button("Cancelar"):
                    st.session_state["confirm_delete"] = False

    else:
        st.info("Nenhuma avaliação para mostrar.")

    st.divider()

    # --------- Adicionar nova avaliação ---------
    with st.expander("Excluir avaliações selecionadas"):
        if IS_STUDENT:
           st.info("Modo aluno: apenas visualização.")
        else:
           confirmar = st.checkbox("Confirmar exclusão")
            if confirmar:
                if st.button("❌ Excluir selecionadas"):
                    avaliacao_db = avaliacao_db[~avaliacao_db["ID"].isin(ids_sel)]
                    _save_avaliacoes_db(avaliacao_db)
                    st.success("Avaliações excluídas.")
                    st.rerun()


            # inputs principais
            r1 = st.columns(4)
            peso = r1[0].number_input("Peso (kg)", min_value=0.0, step=0.1, value=0.0)
            altura = r1[1].number_input("Altura (m)", min_value=0.0, step=0.01, value=0.0)
            idade = r1[2].number_input("Idade", min_value=0, step=1, value=0)
            obs = r1[3].text_input("Observações", value="") if "Observacoes" in base_cols else ""

            st.markdown("**Dobras cutâneas (mm)**")
            d1 = st.columns(5)
            d_pe = d1[0].number_input("D PE (peitoral)", min_value=0.0, step=0.1, value=0.0)
            d_ab = d1[1].number_input("D AB (abdominal)", min_value=0.0, step=0.1, value=0.0)
            d_cx = d1[2].number_input("D CX (coxa)", min_value=0.0, step=0.1, value=0.0)
            d_si = d1[3].number_input("D SI (supra-ilíaca)", min_value=0.0, step=0.1, value=0.0)
            d_tr = d1[4].number_input("D TR (tríceps)", min_value=0.0, step=0.1, value=0.0)

            st.markdown("**Circunferências (cm)**")
            ccs = st.columns(4)
            cc = ccs[0].number_input("CC (cintura)", min_value=0.0, step=0.1, value=0.0)
            cq = ccs[1].number_input("CQ (quadril)", min_value=0.0, step=0.1, value=0.0)
            ca = ccs[2].number_input("CA (abdômen)", min_value=0.0, step=0.1, value=0.0)
            # RCQ calculado, mas deixo visível
            rcq_manual = ccs[3].number_input("RCQ (auto)", min_value=0.0, step=0.0001, value=0.0, disabled=True)

            submitted = st.form_submit_button("Salvar avaliação", disabled=IS_STUDENT)

        if submitted:
            if not nome_novo.strip():
                st.error("Preencha o Nome.")
            elif sexo_val not in {"Homem", "Mulher"}:
                st.error("Selecione o Sexo.")
            else:
                row = {c: None for c in base_cols}
                row["Nome"] = nome_novo.strip()
                row["Data"] = data_nova
                row["Sexo"] = sexo_val

                row["Peso"] = _to_float_or_none(peso, treat_zero_as_none=True)
                row["Altura"] = _to_float_or_none(altura, treat_zero_as_none=True)
                row["Idade"] = _to_float_or_none(idade, treat_zero_as_none=True)

                row["D PE"] = _to_float_or_none(d_pe, treat_zero_as_none=True)
                row["D AB"] = _to_float_or_none(d_ab, treat_zero_as_none=True)
                row["D CX"] = _to_float_or_none(d_cx, treat_zero_as_none=True)
                row["D SI"] = _to_float_or_none(d_si, treat_zero_as_none=True)
                row["D TR"] = _to_float_or_none(d_tr, treat_zero_as_none=True)

                row["CC"] = _to_float_or_none(cc, treat_zero_as_none=True)
                row["CQ"] = _to_float_or_none(cq, treat_zero_as_none=True)
                row["CA"] = _to_float_or_none(ca, treat_zero_as_none=True)

                if "Observacoes" in base_cols:
                    row["Observacoes"] = obs

                row = _recompute_derived(row, base_cols)

                # ID
                row["ID"] = str(abs(hash(f"{row['Nome']}|{row['Data']}|{os.urandom(6).hex()}")))

                db = avaliacao_db.copy()
                db = pd.concat([db, pd.DataFrame([row])], ignore_index=True)
                db = _ensure_id(db)
                _save_avaliacoes_db(db)

                st.success("Avaliação salva.")
                st.rerun()

    st.divider()

    # --------- Editar avaliação ---------
    with st.expander("Editar avaliação"):
        if av.empty:
            st.info("Nenhuma avaliação no filtro atual para editar.")
        else:
            # selecionar por ID
            options = av[["ID", "Data"]].copy()
            options["label"] = options["Data"].astype(str) + " | ID " + options["ID"].astype(str)
            label_to_id = dict(zip(options["label"], options["ID"]))
            sel_label = st.selectbox("Selecione a avaliação", list(label_to_id.keys()))
            sel_id = str(label_to_id[sel_label])

            base_cols = avaliacao_db.columns.tolist()
            row0 = avaliacao_db[avaliacao_db["ID"].astype(str) == sel_id].iloc[0].to_dict()

            with st.form("form_edit_av", border=True):
                c1, c2, c3 = st.columns([2, 1, 1])
                c1.text_input("Nome", value=str(row0.get("Nome", "")), key="edit_nome")
                c2.date_input("Data", value=pd.to_datetime(row0.get("Data"), errors="coerce").date(), key="edit_data")
                c3.selectbox("Sexo", ["Homem", "Mulher"], index=0 if row0.get("Sexo") == "Homem" else 1, key="edit_sexo")

                r1 = st.columns(4)
                r1[0].number_input("Peso (kg)", value=float(pd.to_numeric(row0.get("Peso"), errors="coerce") or 0.0), min_value=0.0, step=0.1, key="edit_peso")
                r1[1].number_input("Altura (m)", value=float(pd.to_numeric(row0.get("Altura"), errors="coerce") or 0.0), min_value=0.0, step=0.01, key="edit_altura")
                r1[2].number_input("Idade", value=int(pd.to_numeric(row0.get("Idade"), errors="coerce") or 0), min_value=0, step=1, key="edit_idade")
                r1[3].text_input("Observações", value=str(row0.get("Observacoes", "") or ""), key="edit_obs")

                st.markdown("**Dobras cutâneas (mm)**")
                d1 = st.columns(5)
                d1[0].number_input("D PE", value=float(pd.to_numeric(row0.get("D PE"), errors="coerce") or 0.0), min_value=0.0, step=0.1, key="edit_dpe")
                d1[1].number_input("D AB", value=float(pd.to_numeric(row0.get("D AB"), errors="coerce") or 0.0), min_value=0.0, step=0.1, key="edit_dab")
                d1[2].number_input("D CX", value=float(pd.to_numeric(row0.get("D CX"), errors="coerce") or 0.0), min_value=0.0, step=0.1, key="edit_dcx")
                d1[3].number_input("D SI", value=float(pd.to_numeric(row0.get("D SI"), errors="coerce") or 0.0), min_value=0.0, step=0.1, key="edit_dsi")
                d1[4].number_input("D TR", value=float(pd.to_numeric(row0.get("D TR"), errors="coerce") or 0.0), min_value=0.0, step=0.1, key="edit_dtr")

                st.markdown("**Circunferências (cm)**")
                ccs = st.columns(3)
                ccs[0].number_input("CC", value=float(pd.to_numeric(row0.get("CC"), errors="coerce") or 0.0), min_value=0.0, step=0.1, key="edit_cc")
                ccs[1].number_input("CQ", value=float(pd.to_numeric(row0.get("CQ"), errors="coerce") or 0.0), min_value=0.0, step=0.1, key="edit_cq")
                ccs[2].number_input("CA", value=float(pd.to_numeric(row0.get("CA"), errors="coerce") or 0.0), min_value=0.0, step=0.1, key="edit_ca")

                ok = st.form_submit_button("Salvar alterações", disabled=IS_STUDENT)

            if ok:
                updated = {c: row0.get(c) for c in base_cols}
                updated["Nome"] = st.session_state["edit_nome"].strip()
                updated["Data"] = st.session_state["edit_data"]
                updated["Sexo"] = st.session_state["edit_sexo"]

                updated["Peso"] = _to_float_or_none(st.session_state["edit_peso"], treat_zero_as_none=True)
                updated["Altura"] = _to_float_or_none(st.session_state["edit_altura"], treat_zero_as_none=True)
                updated["Idade"] = _to_float_or_none(st.session_state["edit_idade"], treat_zero_as_none=True)

                updated["D PE"] = _to_float_or_none(st.session_state["edit_dpe"], treat_zero_as_none=True)
                updated["D AB"] = _to_float_or_none(st.session_state["edit_dab"], treat_zero_as_none=True)
                updated["D CX"] = _to_float_or_none(st.session_state["edit_dcx"], treat_zero_as_none=True)
                updated["D SI"] = _to_float_or_none(st.session_state["edit_dsi"], treat_zero_as_none=True)
                updated["D TR"] = _to_float_or_none(st.session_state["edit_dtr"], treat_zero_as_none=True)

                updated["CC"] = _to_float_or_none(st.session_state["edit_cc"], treat_zero_as_none=True)
                updated["CQ"] = _to_float_or_none(st.session_state["edit_cq"], treat_zero_as_none=True)
                updated["CA"] = _to_float_or_none(st.session_state["edit_ca"], treat_zero_as_none=True)

                if "Observacoes" in base_cols:
                    updated["Observacoes"] = st.session_state["edit_obs"]

                updated = _recompute_derived(updated, base_cols)
                updated["ID"] = sel_id

                db = avaliacao_db.copy()
                db.loc[db["ID"].astype(str) == sel_id, base_cols] = pd.DataFrame([updated])[base_cols].values
                _save_avaliacoes_db(db)
                st.success("Avaliação atualizada.")
                st.rerun()

    st.divider()

    # --------- Comparar duas avaliações ---------
    with st.expander("Comparar avaliações"):
        if av.shape[0] < 2:
            st.info("Precisa de pelo menos 2 avaliações no filtro atual.")
        else:
            labels = av["Data"].astype(str) + " | " + av["ID"].astype(str)
            map_label_id = dict(zip(labels, av["ID"].astype(str)))
            l1 = st.selectbox("Avaliação A", list(map_label_id.keys()), index=0)
            l2 = st.selectbox("Avaliação B", list(map_label_id.keys()), index=min(1, len(map_label_id)-1))
            id_a, id_b = map_label_id[l1], map_label_id[l2]

            a = avaliacao_db[avaliacao_db["ID"].astype(str) == str(id_a)].iloc[0]
            b = avaliacao_db[avaliacao_db["ID"].astype(str) == str(id_b)].iloc[0]

            def getnum(s):
                return pd.to_numeric(s, errors="coerce")

            campos = ["Peso", "G", "MM", "IMC", "CC", "CQ", "CA", "RCQ"]
            rows = []
            for c in campos:
                if c in avaliacao_db.columns:
                    va = float(getnum(a.get(c))) if pd.notna(getnum(a.get(c))) else None
                    vb = float(getnum(b.get(c))) if pd.notna(getnum(b.get(c))) else None
                    delta = (vb - va) if (va is not None and vb is not None) else None
                    rows.append({"Métrica": c, "A": va, "B": vb, "Diferença (B-A)": delta})

            df_cmp = pd.DataFrame(rows)
            st.dataframe(df_cmp, use_container_width=True, hide_index=True)

# -----------------
# Ficha de treino
# -----------------
if page == "Ficha de treino":
    st.subheader("Ficha de treino")

    if dados_treinos.empty:
        st.info("A aba DADOS_TREINOS não foi encontrada ou está vazia.")
    else:
        # A planilha pode vir em dois formatos:
        # (A) "longo": colunas [Grupo, Exercicio]
        # (B) "wide" (seu caso): cada COLUNA eh um grupo muscular, e as CELULAS sao os exercicios.
        import unicodedata

        def _norm_col(x: object) -> str:
            s = str(x).strip().lower().replace("_", " ")
            s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
            s = " ".join(s.split())
            return s

        grp_col = None
        ex_col = None
        for c in dados_treinos.columns:
            nc = _norm_col(c)
            if nc in {"grupo", "grupo muscular", "grupomuscular"}:
                grp_col = c
            if nc in {"exercicio", "exercicios", "exercicio(s)", "exercicio(s)"}:
                ex_col = c

        is_long_format = grp_col is not None and ex_col is not None

        if is_long_format:
            grupos = sorted([x for x in dados_treinos[grp_col].dropna().astype(str).unique().tolist() if str(x).strip()])
        else:
            # formato wide: os grupos sao os nomes das colunas com pelo menos 1 valor
            grupos = [c for c in dados_treinos.columns if dados_treinos[c].notna().any()]
            grupos = sorted([str(g).strip() for g in grupos if str(g).strip()])

        if "ficha" not in st.session_state:
            st.session_state["ficha"] = []

        # -----------------
        # Planos prontos (simples)
        # -----------------
        with st.expander("Gerar treino pronto (modelos)"):
            st.caption("Gera uma ficha automaticamente com base nos grupos disponíveis na planilha de treinos.")

            modelos = {
                "ABC (Hipertrofia)": {
                    "A": ["PEITO", "TRICEPS", "ABDÔMEN"],
                    "B": ["COSTAS", "BÍCEPS"],
                    "C": ["PERNAS", "GLÚTEOS", "PANTURRILHAS"],
                },
                "Full Body": {
                    "Único": ["PEITO", "COSTAS", "OMBROS", "PERNAS", "ABDÔMEN"],
                },
                "PPL": {
                    "Push": ["PEITO", "OMBROS", "TRICEPS"],
                    "Pull": ["COSTAS", "BÍCEPS"],
                    "Legs": ["PERNAS", "GLÚTEOS", "PANTURRILHAS"],
                },
            }

            modelo = st.selectbox("Modelo", list(modelos.keys()))
            dia = st.selectbox("Dia", list(modelos[modelo].keys()))
            n_ex_por_grupo = st.slider("Exercícios por grupo", 1, 6, 3)
            series_padrao = st.number_input("Séries padrão", 1, 10, 3)
            reps_padrao = st.number_input("Reps padrão", 1, 30, 10)

            def _get_exs_for_group(group_name: str) -> list[str]:
                # tenta casar ignorando acentos/caixa
                def _norm(s: str) -> str:
                    ss = str(s).strip().lower().replace("_", " ")
                    ss = "".join(ch for ch in unicodedata.normalize("NFKD", ss) if not unicodedata.combining(ch))
                    ss = " ".join(ss.split())
                    return ss

                target = _norm(group_name)

                if is_long_format:
                    # encontra o grupo mais parecido
                    poss = dados_treinos[grp_col].dropna().astype(str).unique().tolist()
                    match = None
                    for g in poss:
                        if _norm(g) == target:
                            match = g
                            break
                    if match is None:
                        return []
                    exs0 = (
                        dados_treinos.loc[dados_treinos[grp_col].astype(str) == str(match), ex_col]
                        .dropna()
                        .astype(str)
                        .map(lambda x: x.strip())
                        .tolist()
                    )
                    return [x for x in exs0 if x]

                # wide: colunas = grupos
                cols = list(dados_treinos.columns)
                match_col = None
                for c in cols:
                    if _norm(c) == target:
                        match_col = c
                        break
                if match_col is None:
                    return []
                exs0 = dados_treinos[match_col].dropna().astype(str).map(lambda x: x.strip()).tolist()
                return [x for x in exs0 if x]

            if st.button("Gerar ficha agora", disabled=IS_STUDENT):
                st.session_state["ficha"] = []
                for g in modelos[modelo][dia]:
                    exs_g = _get_exs_for_group(g)
                    for ex in exs_g[:n_ex_por_grupo]:
                        st.session_state["ficha"].append(
                            {
                                "Data": date.today(),
                                "Nome": nome_sel if "nome_sel" in locals() else "(sem nomes)",
                                "Grupo muscular": g,
                                "Exercicio": ex,
                                "Series": int(series_padrao),
                                "Repeticoes": int(reps_padrao),
                                "Carga (kg)": 0.0,
                                "Observacoes": "",
                            }
                        )
                st.success("Ficha gerada. Ajuste cargas/observações e salve no registro.")
                st.rerun()

        c1, c2, c3, c4, c5, c6 = st.columns([1.2, 1.2, 2, 0.8, 0.8, 0.8])
        with c1:
            data_treino = st.date_input("Data", value=date.today())
        with c2:
            nome_treino = st.selectbox("Nome", nomes if nomes else ["(sem nomes)"])
        with c3:
            grupo_sel = st.selectbox("Grupo muscular", grupos)
        if is_long_format:
            exs = (
                dados_treinos.loc[dados_treinos[grp_col].astype(str) == str(grupo_sel), ex_col]
                .dropna()
                .astype(str)
                .map(lambda x: x.strip())
                .unique()
                .tolist()
            )
        else:
            # wide: pega os valores nao nulos da coluna selecionada
            col = str(grupo_sel)
            exs = (
                dados_treinos[col]
                .dropna()
                .astype(str)
                .map(lambda x: x.strip())
                .unique()
                .tolist()
            )
        exs = sorted([x for x in exs if x])
        with c4:
            exercicio_sel = st.selectbox("Exercício", exs if exs else ["(sem exercícios)"])
        with c5:
            series = st.number_input("Séries", min_value=1, step=1, value=3)
        with c6:
            reps = st.number_input("Reps", min_value=1, step=1, value=10)

        carga = st.number_input("Carga (kg)", min_value=0.0, step=0.5, value=0.0)
        obs = st.text_input("Observações", value="")

        add = st.button("Adicionar na ficha", disabled=IS_STUDENT)
        if add:
            st.session_state["ficha"].append(
                {
                    "Data": data_treino,
                    "Nome": nome_treino,
                    "Grupo muscular": grupo_sel,
                    "Exercicio": exercicio_sel,
                    "Series": series,
                    "Repeticoes": reps,
                    "Carga (kg)": carga,
                    "Observacoes": obs,
                }
            )

        df_ficha = pd.DataFrame(st.session_state["ficha"]) if st.session_state["ficha"] else pd.DataFrame()

        st.divider()
        st.markdown("#### Ficha (itens adicionados)")
        if df_ficha.empty:
            st.info("Adicione exercícios para montar a ficha.")
        else:
            st.dataframe(df_ficha, use_container_width=True, hide_index=True)

            c1, c2, c3 = st.columns([1.2, 2, 1.2])
            with c1:
                if st.button("Salvar no registro", disabled=IS_STUDENT):
                    rows = df_ficha.to_dict(orient="records")
                    _append_registro(rows)
                    st.success("Treino salvo no registro.")

            with c2:
                modo_exp = st.radio(
                    "Exportação",
                    ["Treino do dia (modelo do Excel)", "Treinos da semana (modelo do Excel)"],
                    horizontal=True,
                )

                # dados para exportação
                df_dia = df_ficha.copy()
                if "Data" in df_dia.columns:
                    df_dia["Data"] = pd.to_datetime(df_dia["Data"], errors="coerce").dt.date
                    df_dia = df_dia[df_dia["Data"] == data_treino]

                # Semana: usa registro + ficha atual (caso não tenha salvado ainda)
                reg_all = _read_registro()
                if not reg_all.empty:
                    if "Data" in reg_all.columns:
                        reg_all["Data"] = pd.to_datetime(reg_all["Data"], errors="coerce").dt.date
                    if "Nome" in reg_all.columns:
                        reg_all = reg_all[reg_all["Nome"].astype(str) == str(nome_treino)]
                df_semana = pd.concat([reg_all, df_ficha], ignore_index=True) if not reg_all.empty else df_ficha.copy()

                xlsx_bytes = uploaded.getvalue() if uploaded else None
                modo_simple = "Dia" if modo_exp.startswith("Treino do dia") else "Semana"
                excel_model = _export_ficha_excel_model(
                    xlsx_bytes=xlsx_bytes,
                    nome=str(nome_treino),
                    modo=modo_simple,
                    data_base=data_treino,
                    df_dia=df_dia,
                    df_semana=df_semana,
                    obs=str(obs or ""),
                )

                fname = (
                    f"ficha_{str(nome_treino).replace(' ', '_')}_{data_treino.strftime('%Y-%m-%d')}.xlsx"
                    if modo_simple == "Dia"
                    else f"ficha_semana_{str(nome_treino).replace(' ', '_')}_{data_treino.strftime('%Y-%m-%d')}.xlsx"
                )

                st.download_button(
                    "Baixar ficha (Excel - modelo)",
                    data=excel_model,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            with c3:
                pdf_bytes = _make_pdf_from_table("Ficha de Treino", df_ficha)
                st.download_button(
                    "Exportar ficha (PDF simples)",
                    data=pdf_bytes,
                    file_name="ficha_treino.pdf",
                    mime="application/pdf",
                )

            if st.button("Limpar ficha"):
                st.session_state["ficha"] = []
                st.rerun()

# -----------------
# Registro de treinos
# -----------------
if page == "Registro de treinos":
    st.subheader("Registro de treinos")

    reg = _read_registro()
    if reg.empty:
        st.info("Nenhum treino registrado ainda.")

    # filtro por nome
    if "Nome" in reg.columns and nomes:
        reg_f = reg[reg["Nome"].astype(str) == str(nome_sel)].copy()
    else:
        reg_f = reg.copy()

    reg_f = reg_f.sort_values("Data")

    st.markdown("#### Editar / excluir")
    st.caption("Selecione uma ou mais linhas para excluir. Para editar, selecione exatamente 1 linha.")

    view = reg_f.copy()
    if "Selecionar" not in view.columns:
        view.insert(0, "Selecionar", False)

    edited = st.data_editor(
        view,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Selecionar": st.column_config.CheckboxColumn("Selecionar", help="Marque para editar/excluir"),
        },
        disabled=[c for c in view.columns if c != "Selecionar"],
        key="reg_editor",
    )

    selected_ids = []
    if "Selecionar" in edited.columns and "ID" in edited.columns:
        selected_ids = edited.loc[edited["Selecionar"] == True, "ID"].astype(str).tolist()

    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("Excluir selecionados", disabled=(IS_STUDENT or (len(selected_ids) == 0))):
            st.session_state["confirm_del_treinos"] = True

    with c2:
        if st.button("Editar selecionado", disabled=(len(selected_ids) != 1)):
            st.session_state["edit_treino_id"] = selected_ids[0]

    with c3:
        # downloads filtrados
        st.download_button(
            "Baixar registro filtrado (CSV)",
            data=reg_f.drop(columns=[c for c in ["Selecionar"] if c in reg_f.columns]).to_csv(index=False).encode("utf-8"),
            file_name="registro_treinos_filtrado.csv",
            mime="text/csv",
        )

    # Confirmacao exclusao
    if st.session_state.get("confirm_del_treinos", False):
        st.warning(f"Tem certeza que deseja excluir {len(selected_ids)} registro(s)?")
        cc1, cc2 = st.columns(2)
        with cc1:
            if st.button("Confirmar exclusao"):
                reg_new = reg[~reg["ID"].astype(str).isin([str(i) for i in selected_ids])].copy()
                _save_registro(reg_new)
                st.session_state["confirm_del_treinos"] = False
                st.success("Registros excluidos.")
                st.rerun()
        with cc2:
            if st.button("Cancelar"):
                st.session_state["confirm_del_treinos"] = False

    # Edicao (1 linha)
    edit_id = st.session_state.get("edit_treino_id", None)
    if edit_id:
        row = reg.loc[reg["ID"].astype(str) == str(edit_id)].iloc[0].to_dict()
        with st.expander("Editar treino selecionado", expanded=True):
            with st.form("form_edit_treino"):
                d = st.date_input("Data", value=row.get("Data") or date.today())
                nome = st.text_input("Nome", value=str(row.get("Nome", "")))
                grupo = st.text_input("Grupo muscular", value=str(row.get("Grupo muscular", "")))
                ex = st.text_input("Exercicio", value=str(row.get("Exercicio", "")))
                s = st.number_input("Series", min_value=0.0, value=float(row.get("Series") or 0.0), step=1.0)
                r = st.number_input("Repeticoes", min_value=0.0, value=float(row.get("Repeticoes") or 0.0), step=1.0)
                kg = st.number_input("Carga (kg)", min_value=0.0, value=float(row.get("Carga (kg)") or 0.0), step=0.5)
                obs = st.text_input("Observacoes", value=str(row.get("Observacoes", "")))
                ok = st.form_submit_button("Salvar alteracoes", disabled=IS_STUDENT)

            if ok:
                reg_new = reg.copy()
                mask = reg_new["ID"].astype(str) == str(edit_id)
                reg_new.loc[mask, "Data"] = d
                reg_new.loc[mask, "Nome"] = nome
                reg_new.loc[mask, "Grupo muscular"] = grupo
                reg_new.loc[mask, "Exercicio"] = ex
                reg_new.loc[mask, "Series"] = s
                reg_new.loc[mask, "Repeticoes"] = r
                reg_new.loc[mask, "Carga (kg)"] = kg
                reg_new.loc[mask, "Observacoes"] = obs
                _save_registro(reg_new)
                st.session_state["edit_treino_id"] = None
                st.success("Treino atualizado.")
                st.rerun()

    st.divider()
    st.markdown("#### Visualizacao")
    st.dataframe(reg_f.drop(columns=[c for c in ["Selecionar"] if c in reg_f.columns]), use_container_width=True, hide_index=True)

    # -----------------
    # Evolucao de carga / PR
    # -----------------
    st.divider()
    st.markdown("#### Evolução de carga por exercício")

    if not reg_f.empty and ("Exercicio" in reg_f.columns) and ("Data" in reg_f.columns):
        ex_opts = sorted(reg_f["Exercicio"].dropna().astype(str).unique().tolist())
        ex_pick = st.selectbox("Exercício", ex_opts)

        evo = reg_f[reg_f["Exercicio"].astype(str) == str(ex_pick)].copy()
        evo["Data"] = pd.to_datetime(evo["Data"], errors="coerce")
        evo = evo.dropna(subset=["Data"]).sort_values("Data")

        # Carga
        if "Carga (kg)" in evo.columns:
            evo["Carga (kg)"] = pd.to_numeric(evo["Carga (kg)"], errors="coerce")
        if "Series" in evo.columns:
            evo["Series"] = pd.to_numeric(evo["Series"], errors="coerce")
        if "Repeticoes" in evo.columns:
            evo["Repeticoes"] = pd.to_numeric(evo["Repeticoes"], errors="coerce")

        # PR
        pr = None
        if "Carga (kg)" in evo.columns and evo["Carga (kg)"].notna().any():
            pr = float(evo["Carga (kg)"].max())
        c1, c2 = st.columns(2)
        with c1:
            st.metric("PR (carga máxima)", f"{pr:.2f} kg" if pr is not None else "—")
        with c2:
            # volume = series * reps * carga
            vol = None
            if all(col in evo.columns for col in ["Series", "Repeticoes", "Carga (kg)"]):
                evo["Volume"] = evo["Series"].fillna(0) * evo["Repeticoes"].fillna(0) * evo["Carga (kg)"].fillna(0)
                vol = float(evo["Volume"].sum()) if evo["Volume"].notna().any() else None
            st.metric("Volume no período", f"{vol:.0f}" if vol is not None else "—")

        if not evo.empty and "Carga (kg)" in evo.columns:
            fig = px.line(evo, x="Data", y="Carga (kg)", markers=True, title=f"Carga ao longo do tempo - {ex_pick}", template=PLOTLY_TEMPLATE)
            st.plotly_chart(fig, use_container_width=True)

        if not evo.empty and "Volume" in evo.columns:
            fig2 = px.bar(evo, x="Data", y="Volume", title=f"Volume por sessão - {ex_pick}", template=PLOTLY_TEMPLATE)
            st.plotly_chart(fig2, use_container_width=True)
    else:
        st.info("Registre treinos para ver a evolução por exercício.")

    # downloads completos
    c1, c2 = st.columns(2)
    with c1:
        st.download_button("Baixar registro (CSV)", data=reg.to_csv(index=False).encode("utf-8"), file_name="registro_treinos.csv", mime="text/csv")
    with c2:
        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            reg.to_excel(writer, index=False, sheet_name="REGISTRO")
        st.download_button(
            "Baixar registro (Excel)",
            data=out.getvalue(),
            file_name="registro_treinos.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
